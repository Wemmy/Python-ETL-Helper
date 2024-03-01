from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
import smtplib
import pandas as pd
from openpyxl import load_workbook

def get_date():
    """
    Returns today's date in 'yyyy-mm-dd' format.
    """
    return datetime.now().strftime('%Y-%m-%d')


def Send_Email(log_file, subject = 'test'):

    # Email configuration
    sender = 'your_email'
    recipients = ["recipients_email"]
    copy_recipients = "cc_email"

    # Send success email
    # subject = "Canteen Sales: Data Update Finished"
    message = f"Please find attachement for details"
    
    # Compose email
    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = ', '.join(recipients)
    msg["Cc"] = copy_recipients
    msg["Subject"] = subject
    msg.attach(MIMEText(message))
        
    # Attach a file
    with open(log_file, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {log_file}",
    )
    msg.attach(part)

    # Send email
    smtp_server = "your_smtp_server"
    smtp_port = 25
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.send_message(msg)


def read_excel_file(file_name, sheet_name, header_row = 0):
    # Load the Excel file
    wb = load_workbook(filename=file_name, read_only=True, data_only=True)

    # Get the sheet
    sheet = wb[sheet_name]

    # Create a pandas DataFrame from the sheet data
    df = pd.DataFrame(sheet.values)

    # Set the headers to the first row of the DataFrame
    header_cells = sheet[header_row]
    headers = [cell.value for cell in header_cells]

    # Create a pandas DataFrame from the sheet data
    rows = sheet.iter_rows(min_row=header_row+1)
    data = [[cell.value for cell in row] for row in rows]
    df = pd.DataFrame(data, columns=headers)

    # Close the workbook
    wb.close()

    return df


def read_excel_table(file_name: str, sheet_name:str,  table_name: str) -> pd.DataFrame:
    wb = load_workbook(file_name, read_only= False, data_only = True) # openpyxl does not have table info if read_only is True; data_only means any functions will pull the last saved value instead of the formula
    sheet = wb[sheet_name] # get the sheet object instead of string
    tbl = sheet.tables[table_name] # get table object instead of string
    tbl_range = tbl.ref #something like 'C4:F9'

    data = sheet[tbl_range] # returns a tuple that contains rows, where each row is a tuple containing cells
    content = [[cell.value for cell in row] for row in data] # loop through those row/cell tuples
    header = content[0] # first row is column headers
    rest = content[1:] # every row that isn't the first is data
    df = pd.DataFrame(rest, columns = header)
    wb.close()
    return df